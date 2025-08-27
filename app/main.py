import logging
import time

from prometheus_client import start_http_server, Counter, Histogram, Gauge
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from logging_loki import LokiHandler

# ================================
# Configuração do Loki
# ================================
LOKI_HOST = "loki"
LOKI_PORT = 3100

logging.basicConfig(level=logging.INFO)
root_logger = logging.getLogger()
root_logger.addHandler(
    LokiHandler(
        url=f"http://{LOKI_HOST}:{LOKI_PORT}/loki/api/v1/push",
        tags={"job": "automation-app"},
        version="1",
    )
)

logger = logging.getLogger(__name__)

# ================================
# Métricas Prometheus
# ================================
automation_runs = Counter(
    "automation_runs_total", "Número de vezes que a automação foi executada"
)
automation_success = Counter(
    "automation_success_total", "Número de execuções bem-sucedidas da automação"
)
automation_errors = Counter(
    "automation_errors_total", "Número de erros na automação"
)
automation_duration = Histogram(
    "automation_run_duration_seconds", "Duração da execução da automação"
)
automation_latency = Histogram(
    "automation_latency_seconds", "Latência fim-a-fim da automação"
)
app_uptime = Gauge(
    "app_uptime_seconds", "Tempo em segundos que a aplicação está rodando"
)
app_availability = Gauge(
    "app_availability", "Disponibilidade da aplicação (1=up, 0=down)"
)

start_time = time.time()


def run_automation():
    """Executa a automação e atualiza métricas Prometheus e logs no Loki."""
    
    try:
    
        wb = load_workbook("contatos.xlsx")
        ws = wb.active

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(
                base_url="https://docs.google.com/forms/d/17y-FzMD9pU5x2BxBikMlLMlkoRKBtlt7viNBoDY4PN4/preview"
            )

            for j in range(1, ws.max_row + 1):
                valores = [
                    ws.cell(row=j, column=i).value
                    for i in range(1, ws.max_column + 1)
                ]
                
                if j != 1:
                    while True:
                        with automation_duration.time():
                            automation_runs.inc()
                            start = time.time()  # início da execução
                            logger.info("🚀 Iniciando execução da automação...")
                            logger.info(f"📄 Processando linha: {valores}")
                            
                            input_selector = 'input[type="text"][class="whsOnd zHQkBf"]'
                            page.goto(
                                "https://docs.google.com/forms/d/17y-FzMD9pU5x2BxBikMlLMlkoRKBtlt7viNBoDY4PN4/preview"
                            )
                            
                            page.wait_for_selector(
                                input_selector, state="visible", timeout=10000
                            )

                            page.get_by_label("Nome").fill(str(valores[1]), force=True)
                            page.get_by_label("E-mail").fill(str(valores[2]), force=True)
                            page.get_by_label("Endereço").fill(str(valores[3]), force=True)
                            page.get_by_label("Número de telefone").fill(
                                str(valores[4]), force=True
                            )
                            page.get_by_label("LinkedIn").fill(str(valores[5]), force=True)

                            
                            logger.info("✅ Automação concluída com sucesso!")
                            automation_success.inc()
                            app_availability.set(1)

                            end = time.time()
                            automation_latency.observe(end - start)
                            
                            time.sleep(10)

                 
    except Exception as e:
        automation_errors.inc()
        logger.error(f"❌ Erro na automação: {e}")


if __name__ == "__main__":
    # Sobe servidor Prometheus
    start_http_server(8000)
    logger.info("📊 Servidor de métricas rodando em http://localhost:8000/metrics")

    while True:
        run_automation()
        app_uptime.set(time.time() - start_time)
        time.sleep(10)
